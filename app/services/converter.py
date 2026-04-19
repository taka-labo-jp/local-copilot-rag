"""ドキュメント変換サービス。

- 既存: markitdown によるMarkdown変換
- 拡張: XLSXはPandas整形テーブル + 埋め込み画像OCRテキストをチャンク化
- 拡張: DOCX/PPTX/PDFも画像OCRを追加し、可能なら図形を含むページOCRを実施
"""
import hashlib
import io
import logging
import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Any

import pandas as pd
import pytesseract
from markitdown import MarkItDown
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from PIL import Image

logger = logging.getLogger(__name__)

_converter = MarkItDown()


def _safe_slug(text: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z_.-]+", "_", text.strip())
    return slug.strip("_") or "source"


def _source_dirname(source_filename: str) -> str:
    digest = hashlib.sha1(source_filename.encode("utf-8")).hexdigest()[:12]
    return f"{_safe_slug(Path(source_filename).stem)}_{digest}"


def _source_root_dir(image_root_dir: str, source_filename: str) -> Path:
    return Path(image_root_dir) / _source_dirname(source_filename)


def delete_extracted_images(image_root_dir: str, source_filename: str) -> None:
    """指定ソース由来の抽出画像ディレクトリを削除する。"""
    target = _source_root_dir(image_root_dir, source_filename)
    if target.exists():
        shutil.rmtree(target, ignore_errors=True)


def convert_to_markdown(file_bytes: bytes, filename: str) -> str:
    """アップロードされたファイルをMarkdown文字列に変換して返す。"""
    suffix = Path(filename).suffix.lower()
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        result = _converter.convert(tmp_path)
        return result.text_content
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def _render_table_chunk(df: pd.DataFrame, start_row: int, end_row: int, sheet_name: str) -> str:
    headers = [str(c).strip() for c in df.columns]
    row_lines: list[str] = []
    for _, row in df.iterrows():
        values = [str(v).strip() for v in row.tolist()]
        row_lines.append(" | ".join(values))

    body = "\n".join(row_lines)
    return (
        f"Sheet: {sheet_name}\n"
        f"Rows: {start_row}-{end_row}\n"
        f"Columns: {' | '.join(headers)}\n"
        f"{body}"
    ).strip()


def _extract_image_bytes(image: Any) -> bytes | None:
    if hasattr(image, "_data"):
        try:
            return image._data()
        except Exception:
            return None

    ref = getattr(image, "ref", None)
    if isinstance(ref, bytes):
        return ref
    if hasattr(ref, "read"):
        try:
            return ref.read()
        except Exception:
            return None
    return None


def _anchor_to_cell(anchor: Any) -> str:
    anchor_from = getattr(anchor, "_from", None)
    if anchor_from is None:
        return ""

    try:
        col = int(anchor_from.col) + 1
        row = int(anchor_from.row) + 1
        return f"{get_column_letter(col)}{row}"
    except Exception:
        return ""


def _convert_xlsx_to_chunks(
    file_bytes: bytes,
    source_filename: str,
    image_root_dir: str,
    excel_rows_per_chunk: int,
    ocr_lang: str,
) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    rows_per_chunk = max(1, excel_rows_per_chunk)

    # 表データはPandasで正規化してN行単位に分割する。
    try:
        excel = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, dtype=str)
        for sheet_name, df in excel.items():
            if df is None or df.empty:
                continue

            normalized = df.dropna(how="all").dropna(axis=1, how="all").fillna("")
            if normalized.empty:
                continue

            normalized.columns = [str(c).strip() if str(c).strip() else f"col_{i+1}" for i, c in enumerate(normalized.columns)]
            total_rows = len(normalized)

            for start in range(0, total_rows, rows_per_chunk):
                end = min(start + rows_per_chunk, total_rows)
                block = normalized.iloc[start:end]
                text = _render_table_chunk(block, start + 1, end, str(sheet_name))
                if not text:
                    continue

                chunks.append(
                    {
                        "text": text,
                        "metadata": {
                            "content_type": "table",
                            "sheet_name": str(sheet_name),
                            "row_range": f"{start + 1}-{end}",
                            "extractor": "pandas",
                        },
                    }
                )
    except Exception as exc:
        logger.exception("Excel table extraction failed: %s", source_filename)
        raise RuntimeError(f"Excel表データ抽出に失敗しました: {exc}") from exc

    # 画像はopenpyxlで抽出して保存し、TesseractでOCRする。
    src_dir = _source_root_dir(image_root_dir, source_filename)
    src_dir.mkdir(parents=True, exist_ok=True)

    wb = None
    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
        for ws in wb.worksheets:
            sheet_images = list(getattr(ws, "_images", []))
            for idx, image in enumerate(sheet_images, start=1):
                image_bytes = _extract_image_bytes(image)
                if not image_bytes:
                    continue

                ext = str(getattr(image, "format", "png") or "png").lower()
                if ext == "jpeg":
                    ext = "jpg"
                if ext not in {"png", "jpg", "bmp", "gif", "tiff", "webp"}:
                    ext = "png"

                sheet_dir = src_dir / _safe_slug(ws.title)
                sheet_dir.mkdir(parents=True, exist_ok=True)
                image_path = sheet_dir / f"image_{idx}.{ext}"
                image_path.write_bytes(image_bytes)

                ocr_status = "empty"
                ocr_text = ""
                try:
                    with Image.open(io.BytesIO(image_bytes)) as pil_image:
                        ocr_text = pytesseract.image_to_string(pil_image, lang=ocr_lang).strip()
                    ocr_status = "success" if ocr_text else "empty"
                except Exception:
                    ocr_status = "failed"
                    logger.warning("OCR failed: %s (%s #%d)", source_filename, ws.title, idx)

                cell_ref = _anchor_to_cell(getattr(image, "anchor", None))
                text = (
                    f"Sheet: {ws.title}\n"
                    f"Image index: {idx}\n"
                    f"Anchor: {cell_ref or '-'}\n"
                    f"OCR status: {ocr_status}\n"
                    f"OCR text:\n{ocr_text if ocr_text else '(no text extracted)'}"
                )

                chunks.append(
                    {
                        "text": text,
                        "metadata": {
                            "content_type": "image",
                            "sheet_name": ws.title,
                            "image_index": idx,
                            "image_path": str(image_path),
                            "cell_ref": cell_ref,
                            "ocr_status": ocr_status,
                            "extractor": "openpyxl+tesseract",
                        },
                    }
                )
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    return chunks


def _ocr_image_bytes(image_bytes: bytes, ocr_lang: str) -> tuple[str, str]:
    ocr_status = "empty"
    ocr_text = ""
    try:
        with Image.open(io.BytesIO(image_bytes)) as pil_image:
            ocr_text = pytesseract.image_to_string(pil_image, lang=ocr_lang).strip()
        ocr_status = "success" if ocr_text else "empty"
    except Exception:
        ocr_status = "failed"
    return ocr_status, ocr_text


def _extract_zip_media_chunks(
    file_bytes: bytes,
    source_filename: str,
    image_root_dir: str,
    ocr_lang: str,
    suffix: str,
) -> list[dict[str, Any]]:
    """DOCX/PPTX/XLSXのzip内 media を抽出してOCRする。"""
    media_prefix_map = {
        ".docx": "word/media/",
        ".pptx": "ppt/media/",
        ".xlsx": "xl/media/",
    }
    media_prefix = media_prefix_map.get(suffix)
    if not media_prefix:
        return []

    src_dir = _source_root_dir(image_root_dir, source_filename) / "embedded_media"
    src_dir.mkdir(parents=True, exist_ok=True)

    chunks: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            members = sorted([n for n in zf.namelist() if n.startswith(media_prefix) and not n.endswith("/")])
            for idx, member in enumerate(members, start=1):
                data = zf.read(member)
                ext = Path(member).suffix.lower().lstrip(".") or "png"
                if ext == "jpeg":
                    ext = "jpg"
                if ext not in {"png", "jpg", "bmp", "gif", "tiff", "webp"}:
                    continue

                image_path = src_dir / f"media_{idx}.{ext}"
                image_path.write_bytes(data)

                ocr_status, ocr_text = _ocr_image_bytes(data, ocr_lang)
                text = (
                    f"File: {source_filename}\n"
                    f"Embedded media index: {idx}\n"
                    f"Source member: {member}\n"
                    f"OCR status: {ocr_status}\n"
                    f"OCR text:\n{ocr_text if ocr_text else '(no text extracted)'}"
                )
                chunks.append(
                    {
                        "text": text,
                        "metadata": {
                            "content_type": "image",
                            "image_index": idx,
                            "image_path": str(image_path),
                            "archive_member": member,
                            "ocr_status": ocr_status,
                            "extractor": "zip-media+tesseract",
                        },
                    }
                )
    except zipfile.BadZipFile:
        return []
    return chunks


def _convert_office_to_pdf_bytes_with_soffice(
    file_bytes: bytes,
    source_filename: str,
    soffice_bin: str,
) -> bytes | None:
    suffix = Path(source_filename).suffix.lower()
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        input_path = tmp / f"input{suffix}"
        input_path.write_bytes(file_bytes)

        cmd = [
            soffice_bin,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(tmp),
            str(input_path),
        ]
        try:
            proc = subprocess.run(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                check=False,
                timeout=120,
            )
        except (FileNotFoundError, subprocess.TimeoutExpired):
            return None

        if proc.returncode != 0:
            logger.warning(
                "soffice conversion failed for %s: %s",
                source_filename,
                proc.stderr.decode("utf-8", errors="ignore")[:400],
            )
            return None

        out_pdf = tmp / f"input.pdf"
        if not out_pdf.exists():
            candidates = sorted(tmp.glob("*.pdf"))
            if not candidates:
                return None
            out_pdf = candidates[0]
        return out_pdf.read_bytes()


def _ocr_pdf_pages_to_chunks(
    pdf_bytes: bytes,
    source_filename: str,
    image_root_dir: str,
    ocr_lang: str,
    content_type: str,
    extractor: str,
    max_pages: int,
) -> list[dict[str, Any]]:
    """PDFページを画像化してOCRし、ページ単位チャンクを返す。"""
    try:
        import pypdfium2 as pdfium
    except Exception:
        logger.warning("pypdfium2 unavailable, skip page OCR: %s", source_filename)
        return []

    src_dir = _source_root_dir(image_root_dir, source_filename) / content_type
    src_dir.mkdir(parents=True, exist_ok=True)

    chunks: list[dict[str, Any]] = []
    pdf = None
    try:
        pdf = pdfium.PdfDocument(pdf_bytes)
        page_count = min(len(pdf), max(1, max_pages))

        for i in range(page_count):
            page = pdf[i]
            page_image = None
            pil_image = None
            image_bytes = b""
            ocr_status = "failed"
            ocr_text = ""

            try:
                page_image = page.render(scale=2.0)
                pil_image = page_image.to_pil()

                with io.BytesIO() as buffer:
                    pil_image.save(buffer, format="PNG")
                    image_bytes = buffer.getvalue()
                    image_path = src_dir / f"page_{i + 1}.png"
                    image_path.write_bytes(image_bytes)

                ocr_status, ocr_text = _ocr_image_bytes(image_bytes, ocr_lang)
                text = (
                    f"File: {source_filename}\n"
                    f"Page: {i + 1}\n"
                    f"OCR status: {ocr_status}\n"
                    f"OCR text:\n{ocr_text if ocr_text else '(no text extracted)'}"
                )
                chunks.append(
                    {
                        "text": text,
                        "metadata": {
                            "content_type": content_type,
                            "page": str(i + 1),
                            "image_path": str(image_path),
                            "ocr_status": ocr_status,
                            "extractor": extractor,
                        },
                    }
                )
            finally:
                try:
                    if pil_image is not None:
                        pil_image.close()
                except Exception:
                    pass
                try:
                    if page_image is not None:
                        page_image.close()
                except Exception:
                    pass
                try:
                    page.close()
                except Exception:
                    pass
    finally:
        if pdf is not None:
            try:
                pdf.close()
            except Exception:
                pass

    return chunks


def convert_to_chunks(
    file_bytes: bytes,
    filename: str,
    image_root_dir: str,
    excel_rows_per_chunk: int = 10,
    ocr_lang: str = "jpn+eng",
    enable_visual_page_ocr: bool = True,
    max_visual_ocr_pages: int = 30,
    soffice_bin: str = "soffice",
) -> list[dict[str, Any]]:
    """ファイルをチャンク配列に変換する。

    Returns:
        list[dict]: 各要素は {"text": str, "metadata": dict} の形式。
    """
    suffix = Path(filename).suffix.lower()
    chunks: list[dict[str, Any]] = []

    if suffix == ".xlsx":
        xlsx_chunks = _convert_xlsx_to_chunks(
            file_bytes=file_bytes,
            source_filename=filename,
            image_root_dir=image_root_dir,
            excel_rows_per_chunk=excel_rows_per_chunk,
            ocr_lang=ocr_lang,
        )
        chunks.extend(xlsx_chunks)

    if suffix in {".docx", ".pptx", ".xlsx"}:
        chunks.extend(
            _extract_zip_media_chunks(
                file_bytes=file_bytes,
                source_filename=filename,
                image_root_dir=image_root_dir,
                ocr_lang=ocr_lang,
                suffix=suffix,
            )
        )

    if suffix == ".pdf":
        chunks.extend(
            _ocr_pdf_pages_to_chunks(
                pdf_bytes=file_bytes,
                source_filename=filename,
                image_root_dir=image_root_dir,
                ocr_lang=ocr_lang,
                content_type="image",
                extractor="pdf-page+tesseract",
                max_pages=max_visual_ocr_pages,
            )
        )

    office_like = {".docx", ".pptx", ".xlsx", ".doc", ".ppt", ".xls"}
    if enable_visual_page_ocr and suffix in office_like:
        converted_pdf = _convert_office_to_pdf_bytes_with_soffice(
            file_bytes=file_bytes,
            source_filename=filename,
            soffice_bin=soffice_bin,
        )
        if converted_pdf:
            # ページ全体OCRで、画像オブジェクト化されない図形/SmartArt/グラフの文字を拾う。
            chunks.extend(
                _ocr_pdf_pages_to_chunks(
                    pdf_bytes=converted_pdf,
                    source_filename=filename,
                    image_root_dir=image_root_dir,
                    ocr_lang=ocr_lang,
                    content_type="visual_page",
                    extractor="soffice-pdf-page+tesseract",
                    max_pages=max_visual_ocr_pages,
                )
            )

    markdown_text = convert_to_markdown(file_bytes, filename)
    markdown_text = markdown_text.strip()
    if markdown_text:
        chunks.insert(
            0,
            {
                "text": markdown_text,
                "metadata": {
                    "content_type": "text",
                    "extractor": "markitdown",
                },
            },
        )

    return chunks
